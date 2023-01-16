import pytest
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font

from jinja2xlsx.api import render
from jinja2xlsx.config import Config
from jinja2xlsx.style import Style
from jinja2xlsx.testing_utils import read_from_test_dir, get_wb_values, get_test_file_path
from jinja2xlsx.utils import width_pixels_to_xlsx_units, height_pixels_to_xlsx_units


def test_xlsx_table_creation_from_html_table() -> None:
    with read_from_test_dir("table.html") as f:
        html_table = f.read()

    wb = render(html_table)
    assert get_wb_values(wb) == [(1, 2), (3, 4)]


def test_xlsx_table_creation_from_html_table_with_merged_cells() -> None:
    with read_from_test_dir("table_with_merged_cells.html") as f:
        html_table = f.read()
        actual_wb = render(html_table)
        actual_wb_values = get_wb_values(actual_wb)

    expected_wb = load_workbook(get_test_file_path("table_with_merged_cells.xlsx"))
    expected_wb_values = get_wb_values(expected_wb)

    assert actual_wb_values == expected_wb_values


def test_xlsx_table_created_from_html_table_has_styles() -> None:
    with read_from_test_dir("table_with_inline_styles.html") as f:
        html_table = f.read()
        actual_wb = render(html_table)
        styled_cell = actual_wb.active.cell(1, 1)

    assert styled_cell.alignment == Alignment(horizontal="center")
    assert styled_cell.border == Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    assert styled_cell.font == Font(bold=True)


def test_xlsx_table_created_from_html_table_has_side_border() -> None:
    with read_from_test_dir("table_with_side_borders.html") as f:
        html_table = f.read()
        actual_wb = render(html_table)
        styled_cell = actual_wb.active.cell(1, 1)

    assert styled_cell.border == Border(bottom=Side("medium"))


def test_xlsx_table_created_from_html_table_with_merged_cell_and_side_border() -> None:
    with read_from_test_dir("table_with_merged_cells_styled.html") as f:
        html_table = f.read()
        actual_wb = render(html_table)

    assert actual_wb.active.cell(row=1, column=1).border.bottom != Side("medium")
    assert actual_wb.active.cell(row=1, column=2).border.bottom != Side("medium")
    assert actual_wb.active.cell(row=2, column=1).border.bottom == Side("medium")
    assert actual_wb.active.cell(row=2, column=2).border.bottom == Side("medium")


def test_xlsx_table_created_from_table_has_column_width() -> None:
    with read_from_test_dir("table_with_colgroup.html") as f:
        html = f.read()
        wb = render(html)
        # wb.save("actual_table_with_colgroup.xlsx")

    assert wb.active.column_dimensions["A"].width == width_pixels_to_xlsx_units(100)
    assert wb.active.column_dimensions["C"].width == width_pixels_to_xlsx_units(200)


def test_xlsx_table_created_from_html_has_row_height() -> None:
    with read_from_test_dir("table_with_row_height.html") as f:
        html = f.read()
        wb = render(html)
        # wb.save("actual_table_with_row_height.xlsx")

    assert wb.active.row_dimensions[1].height == height_pixels_to_xlsx_units(100)
    assert wb.active.row_dimensions[2].height == height_pixels_to_xlsx_units(200)


def test_xlsx_table_creation_with_default_style() -> None:
    default_style = Style(font=Font("Times New Roman", 15))

    with read_from_test_dir("table.html") as f:
        html_table = f.read()
        wb = render(html_table, default_style=default_style)

    assert wb.active.cell(1, 1).font == default_style.font


def test_xlsx_table_creation_with_wrapping() -> None:
    with read_from_test_dir("table_with_wrap.html") as f:
        html_table = f.read()
        wb = render(html_table)

    assert not wb.active.cell(row=1, column=1).alignment.wrap_text
    assert wb.active.cell(row=1, column=2).alignment.wrap_text


def test_xlsx_creation_from_table_with_multiple_borders() -> None:
    with read_from_test_dir("table_with_multiple_borders.html") as f:
        html_table = f.read()
        wb = render(html_table)

    assert wb.active.cell(1, 1).border.left == Side("thin")
    assert wb.active.cell(1, 1).border.right == Side("thin")
    assert wb.active.cell(1, 1).border.top == Side("thin")


def test_xlsx_creation_from_table_with_image() -> None:
    with read_from_test_dir("table_with_image.html") as f:
        html_table = f.read()
        wb = render(html_table, config=Config(parse_img=True))

    assert len(wb.active._images) == 1


def test_xlsx_creation_from_table_with_image_url() -> None:
    with read_from_test_dir("table_with_image_url.html") as f:
        html_table = f.read()
        wb = render(html_table, config=Config(parse_img=True, parse_img_url=True))

    assert len(wb.active._images) == 1


def test_xlsx_creation_from_table_with_relative_image() -> None:
    with read_from_test_dir("table_with_relative_image.html") as f:
        html_table = f.read()
        wb = render(
            html_table,
            config=Config(
                parse_img=True, parse_img_url=True, base_url="https://avatars.mds.yandex.net"
            ),
        )

    assert len(wb.active._images) == 1


def test_xlsx_creation_from_table_with_relative_image_and_no_base_url() -> None:
    with read_from_test_dir("table_with_relative_image.html") as f:
        html_table = f.read()

    with pytest.raises(ValueError):
        render(html_table, config=Config(parse_img=True, parse_img_url=True))


def test_xlsx_creation_from_table_with_th() -> None:
    with read_from_test_dir("table_with_th.html") as f:
        html_table = f.read()
        wb = render(html_table)

    assert wb.active.cell(row=1, column=1).font.bold
    assert wb.active.cell(row=1, column=1).value == "Номер заказа"
