import os

import pytest
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font

from jinja2xlsx.api import render
from jinja2xlsx.testing_utils import read_from_test_dir, get_wb_values, get_test_file_path


def test_xlsx_table_creation_from_html_table() -> None:
    with read_from_test_dir("table.html") as f:
        html_table = f.read()

    wb = render(html_table)
    assert get_wb_values(wb) == [(1, 2), (3, 4)]


def test_xlsx_table_creation_from_html_table_with_merged_cells() -> None:
    with read_from_test_dir("table_with_merged_cells.html") as f:
        html_table = f.read()
        actual_wb = render(html_table)
        actual_wb_values = get_wb_values(actual_wb)

    expected_wb = load_workbook(get_test_file_path("table_with_merged_cells.xlsx"))
    expected_wb_values = get_wb_values(expected_wb)

    assert actual_wb_values == expected_wb_values


def test_xlsx_table_created_from_html_table_has_styles() -> None:
    with read_from_test_dir("table_with_inline_styles.html") as f:
        html_table = f.read()
        actual_wb = render(html_table)
        styled_cell = actual_wb.active.cell(1, 1)

    assert styled_cell.alignment == Alignment(horizontal="center")
    assert styled_cell.border == Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    assert styled_cell.font == Font(bold=True)


def test_xlsx_table_created_from_html_table_has_side_border() -> None:
    with read_from_test_dir("table_with_side_borders.html") as f:
        html_table = f.read()
        actual_wb = render(html_table)
        styled_cell = actual_wb.active.cell(1, 1)

    assert styled_cell.border == Border(bottom=Side("medium"))


def test_xlsx_table_created_from_html_table_with_merged_cell_and_side_border() -> None:
    with read_from_test_dir("table_with_merged_cells_styled.html") as f:
        html_table = f.read()
        actual_wb = render(html_table)

    assert actual_wb.active.cell(row=1, column=1).border != Border(bottom=Side("medium"))
    assert actual_wb.active.cell(row=1, column=2).border != Border(bottom=Side("medium"))
    assert actual_wb.active.cell(row=2, column=1).border == Border(bottom=Side("medium"))
    assert actual_wb.active.cell(row=2, column=2).border == Border(bottom=Side("medium"))


@pytest.mark.skipif(
    not os.path.exists(get_test_file_path("report.html")),
    reason="No report.html/xlsx present in test_data/",
)
def test_xlsx_report_creation() -> None:
    with read_from_test_dir("report.html") as f:
        html = f.read()
        wb = render(html)
        wb.save("actual_report.xlsx")
        actual_values = get_wb_values(wb)

    expected_wb = load_workbook(get_test_file_path("report.xlsx"))
    expected_values = get_wb_values(expected_wb)

    assert actual_values == expected_values
