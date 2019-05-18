import re
from typing import Dict, Optional, Any, Tuple, Iterable

from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from requests_html import Element

CellRange = Tuple[Tuple[Cell]]
CellGenerator = Iterable[Tuple[Element, Optional[Cell], Optional[CellRange]]]


def union_dicts(dict_1: Dict, dict_2: Dict, with_none_drop: bool = True) -> Dict:
    """
    >>> union_dicts({"a": 1}, {"a": None, "b": 2})
    {'a': 1, 'b': 2}
    """
    if with_none_drop:
        new_dict_2 = {key: value for key, value in dict_2.items() if value is not None}
    else:
        new_dict_2 = dict_2

    return {**dict_1, **new_dict_2}


def try_extract_pixels(pixel_str: Optional[str]) -> Optional[float]:
    """
    >>> try_extract_pixels("100px")
    100.0
    >>> try_extract_pixels("") is None
    True
    """
    if not pixel_str:
        return None

    return float(re.findall(r"(\d+)px", pixel_str)[0])


def parse_cell_value(cell_text: str) -> Any:
    """
    >>> parse_cell_value("") is None
    True
    >>> parse_cell_value("ass")
    'ass'
    >>> parse_cell_value("1")
    1
    >>> parse_cell_value("1.2")
    1.2
    """

    if cell_text == "":
        return None

    if cell_text.isdigit():
        return int(cell_text)

    # float(str) break python for very long non-float string ~_~
    if cell_text.replace(".", "", 1).isdigit():
        return float(cell_text)

    return cell_text


def create_cell_range_str(col_index: int, colspan: int, row_index: int, rowspan: int) -> str:
    start_column = get_column_letter(col_index + 1)
    start_row = row_index + 1
    end_column = get_column_letter(col_index + colspan)
    end_row = row_index + rowspan
    cell_range = f"{start_column}{start_row}:{end_column}{end_row}"
    return cell_range


def width_pixels_to_xlsx_units(pixels: float) -> float:
    return pixels / 7.5


def height_pixels_to_xlsx_units(pixels: float) -> float:
    return pixels * 3 / 4
