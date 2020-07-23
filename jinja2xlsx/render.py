from dataclasses import dataclass
from typing import Tuple

from openpyxl import Workbook
from openpyxl.cell import MergedCell, Cell
from openpyxl.worksheet.worksheet import Worksheet

from jinja2xlsx.adjust import Adjuster
from jinja2xlsx.image import parse_img
from jinja2xlsx.parse import Parser
from jinja2xlsx.style import Stylist
from jinja2xlsx.utils import CellGenerator, create_cell_range_str, parse_cell_value


@dataclass
class Renderer:
    parser: Parser
    stylist: Stylist
    workbook: Workbook = None
    sheet: Worksheet = None

    def __call__(self) -> Workbook:
        self.wb = Workbook()
        self.sheet = self.wb.active

        cells = list(self._generate_cells())
        self._fill_cells(cells)
        self._style_cells(cells)

        adjuster = Adjuster(self.sheet)
        adjuster.adjust_columns(self.parser.columns)
        adjuster.adjust_rows(self.parser.rows)

        return self.wb

    def _generate_cells(self) -> CellGenerator:
        for row_index, row in enumerate(self.parser.rows):
            col_index = 0

            html_cells = [*row.find("th"), *row.find("td")]
            for html_cell in html_cells:
                target_cell, col_index = self._find_free_cell(col_index, row_index)

                colspan = int(html_cell.attrs.get("colspan", 1))
                rowspan = int(html_cell.attrs.get("rowspan", 1))

                if colspan > 1 or rowspan > 1:
                    cell_range_str = create_cell_range_str(col_index, colspan, row_index, rowspan)
                    self.sheet.merge_cells(cell_range_str)
                    yield html_cell, None, self.sheet[cell_range_str]
                else:
                    yield html_cell, target_cell, None

                col_index += colspan

    def _find_free_cell(self, col_index: int, row_index: int) -> Tuple[Cell, int]:
        target_cell = self.sheet.cell(row_index + 1, col_index + 1)

        while True:
            if isinstance(target_cell, MergedCell):
                col_index += 1
                target_cell = self.sheet.cell(row_index + 1, col_index + 1)
            else:
                break

        return target_cell, col_index

    def _fill_cells(self, cells: CellGenerator) -> None:
        for html_cell, cell, cell_range in cells:
            target_cell = None
            if cell:
                target_cell = cell
            if cell_range:
                target_cell = cell_range[0][0]
            assert target_cell

            image_tag = html_cell.find("img", first=True)
            if image_tag:
                image = parse_img(image_tag)
                self.sheet.add_image(image, target_cell.coordinate)
            else:
                target_cell.value = parse_cell_value(html_cell.text)

    def _style_cells(self, cells: CellGenerator) -> None:
        for html_cell, cell, cell_range in cells:
            style = self.stylist.build_style_from_html(html_cell)

            if cell:
                self.stylist.style_single_cell(cell, style)
            elif cell_range:
                self.stylist.style_merged_cells(cell_range, style)
